
from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file, flash
import sqlite3
import json
from datetime import datetime
import pandas as pd
import os
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.protection import SheetProtection

app = Flask(__name__)
app.secret_key = 'carwash_peña_blanca_secret_key'

# Add flash messages context processor
@app.context_processor
def utility_processor():
    return dict(get_flashed_messages=flash)

# Initialize database
def init_db():
    conn = sqlite3.connect('carwash.db')
    cursor = conn.cursor()
    
    # Products table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            description TEXT,
            price REAL NOT NULL,
            stock INTEGER DEFAULT 0,
            category TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Customers table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS customers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            phone TEXT,
            email TEXT,
            address TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Sales table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS sales (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_id INTEGER,
            total REAL NOT NULL,
            tax REAL DEFAULT 0,
            discount REAL DEFAULT 0,
            payment_method TEXT,
            status TEXT DEFAULT 'completed',
            invoice_number TEXT UNIQUE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (customer_id) REFERENCES customers (id)
        )
    ''')
    
    # Sale items table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS sale_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sale_id INTEGER,
            product_id INTEGER,
            quantity INTEGER NOT NULL,
            unit_price REAL NOT NULL,
            subtotal REAL NOT NULL,
            FOREIGN KEY (sale_id) REFERENCES sales (id),
            FOREIGN KEY (product_id) REFERENCES products (id)
        )
    ''')
    
    # Insert sample data
    cursor.execute("SELECT COUNT(*) FROM products")
    if cursor.fetchone()[0] == 0:
        sample_products = [
            ('Lavado Básico', 'Lavado exterior básico', 15000, 999, 'Servicios'),
            ('Lavado Premium', 'Lavado completo interior y exterior', 25000, 999, 'Servicios'),
            ('Encerado', 'Aplicación de cera protectora', 10000, 999, 'Servicios'),
            ('Shampoo para Auto', 'Shampoo especial para vehículos', 8000, 50, 'Productos'),
            ('Cera Líquida', 'Cera protectora líquida', 12000, 30, 'Productos'),
            ('Toalla Microfibra', 'Toalla de secado microfibra', 5000, 25, 'Productos')
        ]
        cursor.executemany("INSERT INTO products (name, description, price, stock, category) VALUES (?, ?, ?, ?, ?)", sample_products)
    
    conn.commit()
    conn.close()

# Database helper functions
def get_db_connection():
    conn = sqlite3.connect('carwash.db')
    conn.row_factory = sqlite3.Row
    return conn

def generate_invoice_number():
    return f"CWP-{datetime.now().strftime('%Y%m%d%H%M%S')}"

def create_protected_excel_inventory():
    """Create a protected Excel file with current inventory"""
    conn = get_db_connection()
    products = conn.execute("SELECT * FROM products ORDER BY category, name").fetchall()
    conn.close()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    
    # Headers
    headers = ['ID', 'Código', 'Nombre', 'Descripción', 'Precio', 'Stock', 'Categoría']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Data
    for row, product in enumerate(products, 2):
        # Generate service codes
        code = generate_product_code(product['name'], product['category'])
        ws.cell(row=row, column=1, value=product['id'])
        ws.cell(row=row, column=2, value=code)
        ws.cell(row=row, column=3, value=product['name'])
        ws.cell(row=row, column=4, value=product['description'])
        ws.cell(row=row, column=5, value=product['price'])
        ws.cell(row=row, column=6, value=product['stock'])
        ws.cell(row=row, column=7, value=product['category'])
    
    # Protect sheet with password
    ws.protection = SheetProtection(password="admin123", sheet=True)
    
    # Save file
    excel_path = "inventario_protegido.xlsx"
    wb.save(excel_path)
    return excel_path

def generate_product_code(name, category):
    """Generate product/service codes"""
    if category == 'Servicios':
        name_lower = name.lower()
        if 'lavado' in name_lower:
            if 'básico' in name_lower or 'basico' in name_lower:
                return 'CWP1'
            elif 'premium' in name_lower or 'completo' in name_lower:
                return 'CWP2'
            else:
                return 'CWP3'
        elif 'encerado' in name_lower or 'cera' in name_lower:
            return 'CWP4'
        elif 'shampoo' in name_lower or 'shampu' in name_lower:
            return 'CWP5'
        elif 'pulida' in name_lower or 'rin' in name_lower:
            return 'CWP6'
        else:
            return 'CWP7'
    else:
        # For products, use first 3 letters + number
        return f"PROD{name[:3].upper()}"

def sync_inventory_from_excel(excel_path, password):
    """Sync inventory from Excel file"""
    try:
        # Load workbook
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Check if sheet is protected
        if ws.protection.sheet:
            # Try to unprotect with password
            try:
                ws.protection.password = password
            except:
                return False, "Contraseña incorrecta"
        
        conn = get_db_connection()
        
        # Clear existing products
        conn.execute("DELETE FROM products")
        
        # Read data from Excel (skip header row)
        for row in range(2, ws.max_row + 1):
            product_id = ws.cell(row=row, column=1).value
            code = ws.cell(row=row, column=2).value
            name = ws.cell(row=row, column=3).value
            description = ws.cell(row=row, column=4).value
            price = ws.cell(row=row, column=5).value
            stock = ws.cell(row=row, column=6).value
            category = ws.cell(row=row, column=7).value
            
            if name and price:  # Basic validation
                conn.execute(
                    "INSERT INTO products (name, description, price, stock, category) VALUES (?, ?, ?, ?, ?)",
                    (name, description or '', price, stock or 0, category or 'Productos')
                )
        
        conn.commit()
        conn.close()
        return True, "Inventario sincronizado exitosamente"
        
    except Exception as e:
        return False, f"Error al sincronizar: {str(e)}"

# Routes
@app.route('/')
def dashboard():
    conn = get_db_connection()
    
    # Get statistics
    total_products = conn.execute("SELECT COUNT(*) FROM products").fetchone()[0]
    total_customers = conn.execute("SELECT COUNT(*) FROM customers").fetchone()[0]
    total_sales = conn.execute("SELECT COUNT(*) FROM sales").fetchone()[0]
    daily_revenue = conn.execute(
        "SELECT COALESCE(SUM(total), 0) FROM sales WHERE DATE(created_at) = DATE('now')"
    ).fetchone()[0]
    
    # Recent sales
    recent_sales = conn.execute("""
        SELECT s.id, s.invoice_number, c.name as customer_name, s.total, s.created_at
        FROM sales s
        LEFT JOIN customers c ON s.customer_id = c.id
        ORDER BY s.created_at DESC
        LIMIT 5
    """).fetchall()
    
    conn.close()
    
    return render_template('dashboard.html', 
                         total_products=total_products,
                         total_customers=total_customers, 
                         total_sales=total_sales,
                         daily_revenue=daily_revenue,
                         recent_sales=recent_sales)

@app.route('/products')
def products():
    conn = get_db_connection()
    products = conn.execute("SELECT * FROM products ORDER BY name").fetchall()
    conn.close()
    return render_template('products.html', products=products)

@app.route('/products/add', methods=['GET', 'POST'])
def add_product():
    if request.method == 'POST':
        name = request.form['name']
        description = request.form['description']
        price = float(request.form['price'])
        stock = int(request.form['stock'])
        category = request.form['category']
        
        conn = get_db_connection()
        conn.execute(
            "INSERT INTO products (name, description, price, stock, category) VALUES (?, ?, ?, ?, ?)",
            (name, description, price, stock, category)
        )
        conn.commit()
        conn.close()
        
        return redirect(url_for('products'))
    
    return render_template('add_product.html')

@app.route('/customers')
def customers():
    conn = get_db_connection()
    customers = conn.execute("SELECT * FROM customers ORDER BY name").fetchall()
    conn.close()
    return render_template('customers.html', customers=customers)

@app.route('/customers/add', methods=['GET', 'POST'])
def add_customer():
    if request.method == 'POST':
        name = request.form['name']
        phone = request.form['phone']
        email = request.form['email']
        address = request.form['address']
        
        conn = get_db_connection()
        conn.execute(
            "INSERT INTO customers (name, phone, email, address) VALUES (?, ?, ?, ?)",
            (name, phone, email, address)
        )
        conn.commit()
        conn.close()
        
        return redirect(url_for('customers'))
    
    return render_template('add_customer.html')

@app.route('/sales')
def sales():
    conn = get_db_connection()
    sales = conn.execute("""
        SELECT s.id, s.invoice_number, c.name as customer_name, s.total, s.created_at, s.status
        FROM sales s
        LEFT JOIN customers c ON s.customer_id = c.id
        ORDER BY s.created_at DESC
    """).fetchall()
    conn.close()
    return render_template('sales.html', sales=sales)

@app.route('/sales/new', methods=['GET', 'POST'])
def new_sale():
    if request.method == 'POST':
        customer_id = request.form.get('customer_id') or None
        payment_method = request.form['payment_method']
        items = json.loads(request.form['items'])
        
        conn = get_db_connection()
        
        # Calculate totals
        total = sum(float(item['subtotal']) for item in items)
        invoice_number = generate_invoice_number()
        
        # Insert sale
        cursor = conn.execute(
            "INSERT INTO sales (customer_id, total, payment_method, invoice_number) VALUES (?, ?, ?, ?)",
            (customer_id, total, payment_method, invoice_number)
        )
        sale_id = cursor.lastrowid
        
        # Insert sale items and update stock
        for item in items:
            conn.execute(
                "INSERT INTO sale_items (sale_id, product_id, quantity, unit_price, subtotal) VALUES (?, ?, ?, ?, ?)",
                (sale_id, item['product_id'], item['quantity'], item['unit_price'], item['subtotal'])
            )
            
            # Update product stock
            conn.execute(
                "UPDATE products SET stock = stock - ? WHERE id = ?",
                (item['quantity'], item['product_id'])
            )
        
        conn.commit()
        conn.close()
        
        return redirect(url_for('view_invoice', sale_id=sale_id))
    
    # GET request
    conn = get_db_connection()
    products = conn.execute("SELECT * FROM products WHERE stock > 0 ORDER BY name").fetchall()
    customers = conn.execute("SELECT * FROM customers ORDER BY name").fetchall()
    conn.close()
    
    return render_template('new_sale.html', products=products, customers=customers)

@app.route('/sales/<int:sale_id>/invoice')
def view_invoice(sale_id):
    conn = get_db_connection()
    
    sale = conn.execute("""
        SELECT s.*, c.name as customer_name, c.phone, c.email, c.address
        FROM sales s
        LEFT JOIN customers c ON s.customer_id = c.id
        WHERE s.id = ?
    """, (sale_id,)).fetchone()
    
    items = conn.execute("""
        SELECT si.*, p.name as product_name
        FROM sale_items si
        JOIN products p ON si.product_id = p.id
        WHERE si.sale_id = ?
    """, (sale_id,)).fetchall()
    
    conn.close()
    
    return render_template('invoice.html', sale=sale, items=items)

@app.route('/reports')
def reports():
    return render_template('reports.html')

@app.route('/reports/sales_excel')
def export_sales_excel():
    conn = get_db_connection()
    
    sales_data = conn.execute("""
        SELECT 
            s.invoice_number as 'Número Factura',
            c.name as 'Cliente',
            s.total as 'Total',
            s.payment_method as 'Método Pago',
            s.created_at as 'Fecha'
        FROM sales s
        LEFT JOIN customers c ON s.customer_id = c.id
        ORDER BY s.created_at DESC
    """).fetchall()
    
    conn.close()
    
    # Convert to DataFrame
    columns = ['Número Factura', 'Cliente', 'Total', 'Método Pago', 'Fecha']
    df = pd.DataFrame(sales_data, columns=columns)
    
    # Create Excel file in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Ventas', index=False)
    
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'ventas_carwash_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )

@app.route('/reports/inventory_excel')
def export_inventory_excel():
    """Export protected Excel inventory file"""
    excel_path = create_protected_excel_inventory()
    
    return send_file(
        excel_path,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'inventario_protegido_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )

@app.route('/inventory/sync', methods=['GET', 'POST'])
def sync_inventory():
    """Sync inventory from uploaded Excel file"""
    if request.method == 'POST':
        password = request.form.get('password')
        
        if not password:
            flash('Debes ingresar la contraseña', 'error')
            return redirect(url_for('sync_inventory'))
        
        # Check if file was uploaded
        if 'excel_file' not in request.files:
            flash('No se seleccionó ningún archivo', 'error')
            return redirect(url_for('sync_inventory'))
        
        file = request.files['excel_file']
        if file.filename == '':
            flash('No se seleccionó ningún archivo', 'error')
            return redirect(url_for('sync_inventory'))
        
        if file and file.filename.endswith('.xlsx'):
            # Save uploaded file temporarily
            temp_path = f"temp_inventory_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
            file.save(temp_path)
            
            # Sync inventory
            success, message = sync_inventory_from_excel(temp_path, password)
            
            # Remove temporary file
            os.remove(temp_path)
            
            if success:
                flash(message, 'success')
                return redirect(url_for('products'))
            else:
                flash(message, 'error')
        else:
            flash('Solo se permiten archivos .xlsx', 'error')
    
    return render_template('sync_inventory.html')

@app.route('/api/search_product')
def search_product():
    """Search product by name or barcode for POS"""
    query = request.args.get('q', '')
    
    conn = get_db_connection()
    products = conn.execute("""
        SELECT id, name, description, price, stock, category 
        FROM products 
        WHERE name LIKE ? OR description LIKE ?
        ORDER BY name
        LIMIT 10
    """, (f'%{query}%', f'%{query}%')).fetchall()
    conn.close()
    
    results = []
    for product in products:
        code = generate_product_code(product['name'], product['category'])
        results.append({
            'id': product['id'],
            'name': product['name'],
            'description': product['description'],
            'price': product['price'],
            'stock': product['stock'],
            'category': product['category'],
            'code': code
        })
    
    return jsonify(results)

if __name__ == '__main__':
    # Create templates directory if it doesn't exist
    os.makedirs('templates', exist_ok=True)
    os.makedirs('static/css', exist_ok=True)
    
    init_db()
    app.run(host='0.0.0.0', port=5000, debug=True)
